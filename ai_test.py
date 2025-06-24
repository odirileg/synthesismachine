import openpyxl
import random

# 📂 Load Excel file
excel_file = "basic_rl.xlsx"
wb = openpyxl.load_workbook(excel_file)
ws = wb["Experiments"]

# 🎯 Target value (AI must match this)
TARGET = 10  

# Possible changes AI can make to R1 and R2
actions = [-1, 0, 1]  # Decrease, Hold, Increase

# Read initial values from Excel (Convert to float)
r1 = float(ws.cell(row=2, column=2).value)  
r2 = float(ws.cell(row=2, column=3).value)  

# 🏆 Track the best mix (lowest error)
best_r1, best_r2 = r1, r2
best_error = abs(TARGET - (r1 + r2))

# 🔄 Run 10 optimization attempts
for attempt in range(1, 11):
    # 🎲 Choose random changes for R1 and R2
    change_r1 = random.choice(actions)
    change_r2 = random.choice(actions)

    # 🔄 Apply changes (Prevent negatives)
    r1 = max(0, r1 + change_r1)  
    r2 = max(0, r2 + change_r2)

    # 📉 Calculate error (distance from TARGET)
    error = abs(TARGET - (r1 + r2))

    # 🏆 Update best result if this is the lowest error so far
    if error < best_error:
        best_r1, best_r2, best_error = r1, r2, error

    # 📝 Write new attempt to Excel
    ws.append([attempt, r1, r2, error])

# ✨ Save the best result in Excel
best_result_row = ["Best Mix", best_r1, best_r2, best_error]
ws.append([""])  # Add an empty row
ws.append(best_result_row)

# 💾 Save changes to Excel
wb.save(excel_file)

# 🔥 Print best result
print("\n✅ Optimization complete!")
print(f"🏆 Best mix: R1 = {best_r1}, R2 = {best_r2}, Error = {best_error}")
print("🔍 Check the Excel file for the results.")
