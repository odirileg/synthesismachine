#from tkinter import filedialog, messagebox
#import pandas as pd  # For handling CSV files
from openpyxl import load_workbook
import random

'''

def upload_csv(self):
        """Open a file dialog for the user to upload a CSV file."""
        try:
            # Open file dialog and allow user to select a CSV file
            file_path = filedialog.askopenfilename(
                title="Select a CSV File",
                filetypes=[("CSV Files", "*.csv")],
            )
            
            if file_path:  # Check if a file was selected
                # Use pandas to read the CSV file
                data = pd.read_csv(file_path)
                print(data.head())  # Display the first few rows (for testing)
                messagebox.showinfo("Upload Successful", f"File '{file_path}' uploaded successfully!")
            else:
                messagebox.showwarning("No File Selected", "Please select a CSV file.")
        
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

'''
filename = "volume_combinations.xlsx"
workbook = load_workbook(filename)
sheet = workbook['Session 1']
vials = int(input("How many vials would you like to use? \n"))
file_length = vials + 2

from openpyxl import Workbook
import random

def generate_excel(filename):
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Reagent Data"

    # Define the column headers
    headers = ['Reagent 1', 'Volume 1', 'NaNO2', 'Reagent 2', 'Volume 2', 'NaOH', 'HEX']
    ws.append(headers)

    # Generate 24 rows of data
    for _ in range(24):
        # Generate Reagent 1 and Reagent 2 values (1-4)
        reagent1 = random.randint(1, 4)
        reagent2 = random.randint(1, 4)

        # Generate Volume 1 (random value between 0.5 and 3.5 with 0.5 steps)
        volume1 = random.choice([round(i, 1) for i in range(5, 36, 5)]) / 10
        # Calculate Volume 2
        volume2 = round(4 - volume1, 1)

        # Fixed values for NaNO2 and NaOH
        na_no2 = 2
        na_oh = 2

        # Empty value for HEX
        hex_value = ''

        # Append the row to the worksheet
        ws.append([reagent1, volume1, na_no2, reagent2, volume2, na_oh, hex_value])

    # Save the workbook to a file
    wb.save(filename)
    print(f"Excel file '{filename}' generated successfully!")

def read_values_from_columns(sheet, file_length):
    """Reads values from multiple columns in a sheet, replacing empty cells with 0."""
    num_columns = 4  # Define the number of columns to read
    volumes = [[] for _ in range(num_columns)]  # Initialize lists for each column

    for col in range(1, num_columns + 1):  # Start from row 2
        for row in range(2, file_length):  # Loop through columns 1 to num_columns
            cell_value = sheet.cell(row=row, column=col).value or 0  # Replace None with 0
            volumes[row - 1].append(cell_value)  # Append to the appropriate list

    return volumes  # Return a list of lists with values from each column

def function1(*args):
    """Processes and prints values passed as arguments one at a time."""
    print("Processing values in function1:")
    
    for value in enumerate(args):
        # Dynamic processing: Print each value
        print(value)  # Print the current value

        # Example processing: you can replace this with your logic
        processed_value = value * 2  # Example calculation
        print(f"Processed value: {processed_value}\n")  # Print the processed value

generate_excel(filename)
tools = read_values_from_columns(sheet, file_length)
