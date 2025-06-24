from openpyxl import Workbook, load_workbook
import random
from calibration import *


# Generate an Excel file with 24 rows and specific column headers

filename = "reagent_data.xlsx"
vials = int(input("How many vials would you like to use? \n"))
file_length = vials + 2
num_columns_to_read = 8
def generate_excel(filename):
    """
    Generate an Excel file with 9 columns and 24 rows.
    Columns:
        - Reagent 1 Name: Reagent name
        - Reagent 1 Number: Assigned reagent number
        - Volume 1: Random values (0.5 to 3.5, step 0.5)
        - NaNO2: Fixed value (2)
        - Reagent 2 Name: Reagent name
        - Reagent 2 Number: Assigned reagent number
        - Volume 2: Computed as (4 - Volume 1)
        - NaOH: Fixed value (2)
        - HEX: Empty
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Reagent Data"

    # Headers
    headers = [
        'Reagent 1 Name', 'Reagent 1 Number', 'Volume 1', 'NaNO2',
        'Reagent 2 Name', 'Reagent 2 Number', 'Volume 2', 'NaOH', 'HEX'
    ]
    ws.append(headers)

    # Reagents list with names
    reagents = ["Aniline", "Dimethylaniline", "Nitroaniline", "Naphthyamine"]

    # Reagent mapping (Name → Number)
    reagent_mapping = {
        "Aniline": 4,
        "Dimethylaniline": 5,
        "Nitroaniline": 6,
        "Naphthyamine": 7
    }

    # Generate 24 rows of random data
    for _ in range(24):
        reagent1 = random.choice(reagents)
        reagent2 = random.choice(reagents)
        volume1 = round(random.uniform(0.5, 3.5), 1)
        volume2 = round(4 - volume1, 1)

        # Append both names and corresponding numbers
        row = [
            reagent1, reagent_mapping[reagent1],  # Reagent 1 (Name, Number)
            volume1, 2,  # Volume 1, NaNO2
            reagent2, reagent_mapping[reagent2],  # Reagent 2 (Name, Number)
            volume2, 2, ''  # Volume 2, NaOH, HEX
        ]
        ws.append(row)

    wb.save(filename)
    print(f"Excel file '{filename}' generated successfully!")

# Read specific columns from an Excel sheet
def read_values_from_columns(sheet, num_columns, file_length):
    """
    Reads values from specified columns in an Excel sheet.
    Replaces empty cells with 0 and returns the data as a list of lists.

    Args:
        sheet (Worksheet): The worksheet object.
        num_columns (int): Number of columns to read.
        file_length (int): Number of rows to read (starting from row 2).

    Returns:
        List[List]: A list of lists with values from each column.
    """
    volumes = [[] for _ in range(num_columns)]

    for col in range(1, num_columns + 1):
        for row in range(2, file_length):
            cell_value = sheet.cell(row=row, column=col).value or 0
            volumes[col - 1].append(cell_value)

    return volumes

#assigning columns to values
def data_change():
    R1_Name = column_data[0]
    R1 = column_data[1]
    V1 = column_data [2]
    V_Nitrite = column_data[3]
    R2_Name = column_data[4]
    R2 = column_data[5]
    V2 = column_data[6]
    V_NaOH = column_data[7]

generate_excel(filename)

workbook = load_workbook(filename)
sheet = workbook["Reagent Data"]

column_data = read_values_from_columns(sheet, num_columns_to_read, file_length)
data_change()


